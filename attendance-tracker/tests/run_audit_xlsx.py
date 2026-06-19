import os
import re
import openpyxl


def run_audit(spreadsheet_id: str):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.abspath(os.path.join(script_dir, "../audit_sheet.xlsx"))
    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} does not exist.")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    expected_headers = [
        "Age Group",
        "Gender",
        "Preferred Name",
        "Last Name",
        "Present",
        "Scratch",
        "Medley Relay",
        "Free Relay",
        "Free",
        "Back",
        "Breast",
        "Fly",
        "IM",
        "ID",
        "First Name",
        "Age",
        "Team",
    ]

    print("=== AUDIT START ===")

    # 1. Main Tab
    if "Main" not in wb.sheetnames:
        print("FAIL: 'Main' tab not found.")
        return

    main_sheet = wb["Main"]

    # a) Headers
    main_headers = [cell.value for cell in main_sheet[1]]
    main_headers = [
        h.strip() if isinstance(h, str) else h
        for h in main_headers[: len(expected_headers)]
    ]
    if main_headers != expected_headers:
        print("FAIL: 'Main' tab headers do not match expected.")
        print(f"  Got: {main_headers}")
        print(f"  Expected: {expected_headers}")
    else:
        print("PASS: 'Main' tab headers match exactly.")

    # Get all swimmers from Main
    swimmers = []
    for r in range(2, main_sheet.max_row + 1):
        row_vals = [
            main_sheet.cell(row=r, column=c).value
            for c in range(1, len(expected_headers) + 1)
        ]
        if not any(row_vals):
            continue
        swimmer = dict(zip(expected_headers, row_vals))
        swimmers.append(swimmer)

    print(f"Loaded {len(swimmers)} swimmers from 'Main' tab.")

    # b) & c) Sorting and Age Group prefix checks
    sorted_correctly = True
    prefix_correct = True

    last_ag = ""
    last_gender = ""
    last_pref_name = ""

    for idx, s in enumerate(swimmers):
        ag = str(s["Age Group"] or "")
        gender = str(s["Gender"] or "")
        pref_name = str(s["Preferred Name"] or "")

        # c) Age Group column has space prefix for '<= 10' age groups and no prefix for others
        clean_ag = ag.strip()
        is_under_10 = clean_ag in ["6 & Under", "7-8", "9-10"]
        has_space = ag.startswith(" ")

        if is_under_10 and not has_space:
            print(
                f"FAIL: Age Group '{ag}' for swimmer {pref_name} {s['Last Name']} is missing leading space."
            )
            prefix_correct = False
        elif not is_under_10 and has_space:
            print(
                f"FAIL: Age Group '{ag}' for swimmer {pref_name} {s['Last Name']} has unexpected leading space."
            )
            prefix_correct = False

        # check sorting: Age Group -> Gender -> Preferred Name (case-insensitive)
        if idx > 0:
            prev_key = (last_ag, last_gender, last_pref_name.lower())
            curr_key = (ag, gender, pref_name.lower())

            if curr_key < prev_key:
                print(
                    f"FAIL: Sorting mismatch at row {idx + 2}: {prev_key} is after {curr_key}"
                )
                sorted_correctly = False

        last_ag = ag
        last_gender = gender
        last_pref_name = pref_name

    if prefix_correct:
        print("PASS: Age Group space prefixes verified correctly.")
    if sorted_correctly:
        print(
            "PASS: Swimmer entries in 'Main' sorted correctly (Age Group -> Gender -> Preferred Name (case-insensitive))."
        )

    # 2. Age Group Tabs
    age_group_tabs = ["6 & Under", "7-8", "9-10", "11-12", "13-14", "15-18"]

    main_by_ag = {}
    for s in swimmers:
        ag_clean = str(s["Age Group"] or "").strip()
        if ag_clean not in main_by_ag:
            main_by_ag[ag_clean] = []
        main_by_ag[ag_clean].append(s)

    for tab_name in age_group_tabs:
        if tab_name not in wb.sheetnames:
            print(f"FAIL: Age Group tab '{tab_name}' not found.")
            continue

        tab = wb[tab_name]
        tab_headers = [cell.value for cell in tab[1]]
        tab_headers = [
            h.strip() if isinstance(h, str) else h
            for h in tab_headers[: len(expected_headers)]
        ]

        if tab_headers != expected_headers:
            print(f"FAIL: Tab '{tab_name}' headers do not match 'Main'.")
        else:
            print(f"PASS: Tab '{tab_name}' headers match 'Main'.")

        tab_swimmers = []
        for r in range(2, tab.max_row + 1):
            row_vals = [
                tab.cell(row=r, column=c).value
                for c in range(1, len(expected_headers) + 1)
            ]
            if not any(row_vals):
                continue
            swimmer = dict(zip(expected_headers, row_vals))
            tab_swimmers.append(swimmer)

        expected_subset = main_by_ag.get(tab_name, [])
        if len(tab_swimmers) != len(expected_subset):
            print(
                f"FAIL: Tab '{tab_name}' swimmer count ({len(tab_swimmers)}) does not match expected ({len(expected_subset)})."
            )
        else:
            tab_ids = {str(s["ID"]) for s in tab_swimmers}
            expected_ids = {str(s["ID"]) for s in expected_subset}
            if tab_ids != expected_ids:
                print(f"FAIL: Tab '{tab_name}' swimmer IDs mismatch.")
            else:
                print(
                    f"PASS: Tab '{tab_name}' contains the correct subset of swimmers ({len(tab_swimmers)} swimmers)."
                )

    # 3. 'All Scratches' and 'Not Checked In' Tabs
    for tab_name in ["All Scratches", "Not Checked In"]:
        if tab_name not in wb.sheetnames:
            print(f"FAIL: Tab '{tab_name}' not found.")
            continue
        tab = wb[tab_name]
        tab_headers = [cell.value for cell in tab[1]]
        tab_headers = [
            h.strip() if isinstance(h, str) else h
            for h in tab_headers[: len(expected_headers)]
        ]
        if tab_headers != expected_headers:
            print(f"FAIL: Tab '{tab_name}' headers do not match 'Main'.")
        else:
            print(f"PASS: Tab '{tab_name}' headers match 'Main' in row 1.")

        formula_cell = tab["A2"]
        formula_val = formula_cell.value

        norm_val = re.sub(r"\s+", "", str(formula_val or "")).lower()
        norm_val = norm_val.replace("1000", "")

        if tab_name == "All Scratches":
            # Core filter: FILTER(Main!A2:Q, Main!F2:F=TRUE)
            # fallback string: "No Scratches"
            has_filter = (
                "filter(main!a2:q" in norm_val
                and "main!f2:f" in norm_val
                and "true" in norm_val
            )
            has_fallback = "noscratches" in norm_val
            is_valid = has_filter and has_fallback
            expected_desc = (
                "FILTER(Main!A2:Q, Main!F2:F=TRUE) with fallback 'No Scratches'"
            )
        else:
            # Core filter: FILTER(Main!A2:Q, (Main!A2:A<>"") * (Main!E2:E=FALSE) * (Main!F2:F=FALSE))
            # fallback string: "All Checked In"
            has_filter = (
                "filter(main!a2:q" in norm_val
                and "main!a2:a" in norm_val
                and "main!e2:e=false" in norm_val
                and "main!f2:f=false" in norm_val
            )
            has_fallback = "allcheckedin" in norm_val
            is_valid = has_filter and has_fallback
            expected_desc = "FILTER(Main!A2:Q, (Main!A2:A<>'') * (Main!E2:E=FALSE) * (Main!F2:F=FALSE)) with fallback 'All Checked In'"

        if is_valid:
            print(
                f"PASS: Tab '{tab_name}' formula in A2 matches expected logic: {expected_desc}"
            )
        else:
            print(f"FAIL: Tab '{tab_name}' formula in A2 mismatch.")
            print(f"  Got: {formula_val}")
            print(f"  Expected logic: {expected_desc}")

    # 4. 'QR Code' Tab
    if "QR Code" not in wb.sheetnames:
        print("FAIL: 'QR Code' tab not found.")
    else:
        tab = wb["QR Code"]
        a2_val = tab["A2"].value
        if a2_val and "api.qrserver.com" in a2_val and spreadsheet_id in a2_val:
            print(
                "PASS: 'QR Code' tab cell A2 formula contains the expected QR server URL and Spreadsheet ID."
            )
            print(f"  Got formula: {a2_val}")
        else:
            print("FAIL: 'QR Code' tab cell A2 formula invalid.")
            print(f"  Got formula: {a2_val}")

    print("=== AUDIT END ===")


if __name__ == "__main__":
    import sys

    # 1. Try command line argument
    spreadsheet_id_arg = sys.argv[1] if len(sys.argv) > 1 else None

    # 2. Try env variable
    if not spreadsheet_id_arg:
        spreadsheet_id_arg = os.getenv("ATTENDANCE_SPREADSHEET_ID")

    # 3. Try fallback to .env file
    if not spreadsheet_id_arg:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        env_path = os.path.abspath(os.path.join(script_dir, "../../.env"))
        if os.path.exists(env_path):
            with open(env_path, "r") as f:
                for line in f:
                    if line.strip().startswith("ATTENDANCE_SPREADSHEET_ID="):
                        spreadsheet_id_arg = (
                            line.split("=", 1)[1].strip().strip('"').strip("'")
                        )
                        break

    if not spreadsheet_id_arg:
        print(
            "Error: Spreadsheet ID must be passed as an argument or defined in ATTENDANCE_SPREADSHEET_ID."
        )
        sys.exit(1)

    run_audit(spreadsheet_id_arg)
